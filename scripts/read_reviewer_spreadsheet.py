"""Read the reviewer tracking spreadsheet to understand findings and structure."""
import sys
import os

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

SPREADSHEET_PATH = r"C:\Users\sbsch\Dropbox\My Documents\ACAHM\Feb 2026 Meeting\VU\260226 VU Sup Rpt Inst-DAcCHM (C) Pre-Accred. REV TB.xlsx"

def read_with_pandas():
    """Read Excel with pandas."""
    xl_file = pd.ExcelFile(SPREADSHEET_PATH)
    print("Sheet names:")
    for sheet in xl_file.sheet_names:
        print(f"  - {sheet}")
    
    print("\n" + "="*80)
    for sheet_name in xl_file.sheet_names:
        print(f"\nSHEET: {sheet_name}")
        print("="*80)
        df = pd.read_excel(SPREADSHEET_PATH, sheet_name=sheet_name, header=None, nrows=50)
        print(df.to_string(max_rows=50, max_cols=15))
        print("\n")

def read_with_openpyxl():
    """Read Excel with openpyxl."""
    wb = openpyxl.load_workbook(SPREADSHEET_PATH, read_only=True, data_only=True)
    print("Sheet names:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
    
    print("\n" + "="*80)
    for sheet_name in wb.sheetnames:
        print(f"\nSHEET: {sheet_name}")
        print("="*80)
        ws = wb[sheet_name]
        # Read first 50 rows, first 20 columns
        for i, row in enumerate(ws.iter_rows(max_row=50, max_col=20, values_only=True), 1):
            row_str = " | ".join([str(cell)[:40] if cell is not None else "" for cell in row[:15]])
            if row_str.strip():  # Only print non-empty rows
                print(f"Row {i}: {row_str}")
        print("\n")

if __name__ == "__main__":
    if HAS_PANDAS:
        read_with_pandas()
    elif HAS_OPENPYXL:
        read_with_openpyxl()
    else:
        print("Error: Neither pandas nor openpyxl is installed.")
        print("Please install one: pip install pandas openpyxl")
        sys.exit(1)
